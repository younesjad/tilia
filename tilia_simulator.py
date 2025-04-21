import xlwings as xw
import pandas as pd
import openpyxl
import streamlit as st
import os, sys
from openpyxl import load_workbook

def safe_update_excel(path, new_data, sheet_name):
    wb = load_workbook(path, data_only=False, keep_vba=True)  # Set keep_vba=True if needed
    ws = wb[sheet_name]

    for idx, row in new_data.iterrows():
        for col_idx, col_name in enumerate(row.index):
            cell = ws.cell(row=idx+2, column=col_idx+1)
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = row[col_name]

    wb.save(path)

# --- Load Excel tables ---
base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))

excel_path = os.path.join(base_path, 'data', 'Tilia calc testing.xlsx')
image_tilia = os.path.join(base_path, 'data', 'tiliap.jpg')
image_skema = os.path.join(base_path, 'data', 'skema.png')

df_input = pd.read_excel(excel_path, sheet_name="User Guide", engine="openpyxl")
def load_output(result):
    if result==1:
      return pd.read_excel(excel_path, sheet_name="results", engine="openpyxl")
    else:
      return pd.read_excel(excel_path, sheet_name="results2", engine="openpyxl")

# Load the Excel workbook for modifying specific cells
wb = openpyxl.load_workbook(excel_path)
sheet = wb["User Guide"]

st.set_page_config(page_title="Tilia Simulator", layout="wide")

# st.markdown("""
#     <style>
#         /* Hide default Streamlit header and footer */
#         #MainMenu {visibility: hidden;}
#         footer {visibility: hidden;}
#         header {visibility: hidden;}
#     </style>
# """, unsafe_allow_html=True)

# --- Add Custom Styling (CSS) ---
st.markdown("""
    <style>
    /* Global page reset */
    html, body, [class*="css"] {
        font-family: 'Segoe UI', 'Roboto', 'Helvetica Neue', sans-serif;
        background-color: #f7f9fb;
        color: #2c3e50;
        margin: 0;
        padding: 0;
    }
    
    .title-container {
        background: linear-gradient(90deg, #0f2027, #203a43, #2c5364);
        box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        padding: 30px 10px;
        border-radius: 12px;
        text-align: center;
        margin-bottom: 20px;
    }

    .title {
        font-size: 36px;
        font-weight: bold;
        color: #ffffff;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        margin: 0;
    }

    .subtitle {
        font-size: 16px;
        color: #d0d0d0;
        margin-top: 5px;
    }

    /* Section Styling */
    .section {
        background-color: #ffffff;
        border-radius: 10px;
        padding: 25px;
        margin: 20px 0;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.05);
    }

    /* Custom Buttons */
    .stButton>button {        
        background: linear-gradient(90deg, #0f2027, #203a43, #2c5364);
        box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        color: white;
        font-size: 16px;
        font-weight: 600;
        padding: 10px 22px;
        border-radius: 6px;
        border: none;
        transition: background-color 0.3s ease;
    }

    .stButton>button:hover {
        background-color: #219150;
    }

    /* Table Styling */
    .streamlit-table {
        font-size: 14px;
        border-collapse: collapse;
        width: 100%;
    }

    .streamlit-table th {
        background-color: #ecf0f1;
        font-weight: bold;
        text-align: left;
        padding: 10px;
    }

    .streamlit-table td {
        padding: 8px;
        border-top: 1px solid #ddd;
    }

    /* Question cell highlight */
    .question-cell {
        background-color: #e8f5e9 !important;
        font-weight: bold;
        color: #2c3e50;
        border-radius: 5px;
        padding: 8px 10px;
    }

    </style>
""", unsafe_allow_html=True)

# --- Main Content ---
col1, col2, spacer = st.columns([1, 1, 5])
with col1:
    st.image(image_skema, width=300)
with col2:
    st.image(image_tilia, width=300)
st.markdown("""
    <div class="title-container">
        <div class="title">Heat Generation Technologies Simulator</div>
    </div>
""", unsafe_allow_html=True)

# --- Input Section ---
st.markdown("""
    <div class="title-container">
        <div class="title">✍️ Inputs:</div>
    </div>
""", unsafe_allow_html=True)

# --- Define per-cell rules ---
dropdown_cells = {(0, 'Input'): ['Île-de-France','Bourgogne-Franche-Comté', 'Bretagne', 'Centre-Val de Loire', 'Corse', 'Grand Est', 'Hauts-de-France', 'Île-de-France', 'Normandie', 'Nouvelle-Aquitaine', 'Occitanie', 'Pays de la Loire'], (2, 'Input'): ['Yes','No'], (4, 'Input'): ['ON','OFF']}
numeric_cells = {(1, 'Input'): None, (3, 'Input'): None}
text_cells = {}

# Create a copy to hold edited data
new_data = df_input.copy()

# --- Build the input form ---
for idx, row in df_input.iterrows():
    row_values = {}

    for col in df_input.columns:
        cell_key = (idx, col)
        current_value = row[col]

        # Dropdown logic
        if cell_key in dropdown_cells:
            options = dropdown_cells[cell_key]
            default_idx = options.index(current_value) if current_value in options else 0
            new_val = st.selectbox(
                label="", options=options, index=default_idx, key=f"{idx}_{col}"
            )

        # Numeric logic (no min/max, just free numeric input)
        elif cell_key in numeric_cells:
            try:
                numeric_default = float(current_value)
            except:
                numeric_default = 0.0

            new_val = st.number_input(
                label="", value=numeric_default, key=f"{idx}_{col}"
            )

        # Default text input
        elif cell_key in text_cells:
            new_val = st.text_input(label="", value=str(current_value), key=f"{idx}_{col}")
        else:
            # Display the 'Question' value in styled markdown
            st.markdown(f'<div style="background-color:#cbd1e5; box-shadow: 0 4px 12px rgba(0,0,0,0.2); padding:10px; border-radius:5px; font-weight:bold;">{current_value}</div>',unsafe_allow_html=True)
            new_val = current_value  # Don't allow editing

        row_values[col] = new_val

    # Update new_data DataFrame
    for col_name, val in row_values.items():
        new_data.at[idx, col_name] = val

    # --- Update specific cell in Excel sheet (keep formulas) ---
    for col_idx, col_name in enumerate(df_input.columns):
        cell = sheet.cell(row=idx+2, column=col_idx+1)  # Excel is 1-indexed
        new_value = row_values[col_name]
        
        # Update the cell value in the sheet, preserving formulas
        if isinstance(new_value, (int, float)):
            cell.value = new_value
        else:
            cell.value = str(new_value)  # For strings, make sure they're saved as strings

# --- Summary Section ---
st.markdown("""
    <div class="title-container">
        <div class="title">✅ Summary of Inputs:</div>
    </div>
""", unsafe_allow_html=True)
# Add custom class for table
st.markdown("<div class='streamlit-table'>", unsafe_allow_html=True)
st.dataframe(new_data, use_container_width=True)
st.markdown("</div>", unsafe_allow_html=True)

# --- Buttons Side-by-Side with Reduced Space ---
col1, spacer, col2 = st.columns([0.4, 0.05, 2])  # Adjust spacer to reduce gap

if st.button("Save Changes"):
    try:
        safe_update_excel(excel_path, new_data, sheet_name="User Guide")
        st.success("Changes saved successfully without overwriting formulas.")
    except Exception as e:
        st.error(f"Error: {e}")

if st.button("Show Results:", key="show_results"):
    df_output1 = load_output(1)
    st.success("📥 Output table loaded from Excel:")
    st.dataframe(df_output1, use_container_width=True)
    df_output2 = load_output(2)
    st.success("📥 Conclusion:")
    st.dataframe(df_output2, use_container_width=True)
